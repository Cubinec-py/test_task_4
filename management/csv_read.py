import csv
import re

from openpyxl import load_workbook
from datetime import datetime
from fastapi import HTTPException


def read_csv(file_route):
    with open(f"{file_route}", "r") as file:
        csv_reader = csv.reader(file, delimiter="\t")
        header = next(csv_reader)
        date_reg = r"^\d{2}\.\d{2}\.\d{4}$"
        data_to_insert = []
        for row in csv_reader:
            row_dict = {}
            for idx, column in enumerate(header):
                if not row[idx]:
                    row[idx] = None
                elif re.match(date_reg, row[idx]):
                    row[idx] = "-".join(row[idx].split(".")[::-1])
                row_dict[column] = row[idx]
            data_to_insert.append(row_dict)
        return data_to_insert


def read_xlsx(file):
    excel_data = []
    wb = load_workbook(file.file)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        if all(row):
            excel_data.append(row)
    formatted_data = convert_to_desired_format(excel_data)
    return formatted_data


def convert_to_desired_format(excel_data):
    headers = excel_data[0]
    data = []
    for row in excel_data[1:]:
        row_dict = {}
        for i, header in enumerate(headers):
            if header.lower() == "period":
                try:
                    value = row[i].date()
                except AttributeError:
                    value = row[i]
                try:
                    date_obj = datetime.strptime(str(value), "%Y-%m-%d")
                except ValueError:
                    raise HTTPException(
                        status_code=404, detail=f"Invalid date format: {value}."
                    )
                if date_obj.day != 1:
                    raise HTTPException(
                        status_code=404,
                        detail=f"Invalid date: {value} is not the first day of the month",
                    )
                row_dict[header] = date_obj.strftime("%Y-%m-%d")
            else:
                row_dict[header] = row[i]
        data.append(row_dict)
    return data
